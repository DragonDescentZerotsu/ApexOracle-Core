#!/usr/bin/env python3
"""Build the corrected 73-row candidate copy and audit structural diversity.

The external guided-generation outputs, legacy peptide parser, and final wet-lab
workbook are read-only inputs. All generated assets are written under one
reviewer experiment directory with source hashes and row-level lineage.
"""

from __future__ import annotations

import argparse
import csv
import importlib
import json
import re
import shutil
import sys
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import selfies as sf
from openpyxl import load_workbook
from rdkit import Chem, DataStructs


REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "src"))

from apexoracle.evaluation.generated_candidate_diversity import (  # noqa: E402
    FINGERPRINT_BITS,
    FINGERPRINT_INCLUDE_CHIRALITY,
    FINGERPRINT_RADIUS,
    REPLACEMENT_RULES,
    all_pairwise_tanimoto,
    canonical_isomeric_smiles,
    histogram_rows,
    match_replacement_rule,
    morgan_fingerprints,
    morgan_fingerprints_from_mols,
    normalize_legacy_arg,
    sample_distinct_ordered_pairs,
    sampled_tanimoto,
    sha256_file,
    summarize_similarities,
    topology_aware_equal,
)


STRICT_PATTERN = re.compile(
    r"^strain_(BAA-\d+)_MIC_1_length_(\d+)_noise\.txt$"
)
EXPECTED_WORKBOOK_SHA256 = (
    "6f9b5b15a9db76fc5df744f9f51524f57f24e09662cc79a0655b3b0d42c80f22"
)
EXPECTED_STRICT_ROWS = 73
EXPECTED_GENERATION_ROWS = 84_226
TARGET_STRAINS = ("BAA-3170", "BAA-3197")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--strict-dir",
        type=Path,
        default=Path(
            "/data2/tianang/projects/discrete-diffusion-guidance/outputs/"
            "generated_mol_SELFIES_w_mic-new"
        ),
    )
    parser.add_argument(
        "--generation-dir",
        type=Path,
        default=Path(
            "/data2/tianang/projects/discrete-diffusion-guidance/outputs/"
            "generated_mol_SELFIES-new"
        ),
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=Path(
            "/data2/tianang/projects/ApexOracle_cleaned/docs/"
            "ApexOracle_Nat_Biotech/ApexOracle_MIC_data/Summary_pathogens.xlsx"
        ),
    )
    parser.add_argument(
        "--legacy-parser-root",
        type=Path,
        default=Path("/data2/tianang/projects/mdlm"),
    )
    parser.add_argument(
        "--peplink-root",
        type=Path,
        default=Path("/data2/tianang/projects/PepLink"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=REPO_ROOT / "experiments" / "generated_candidate_diversity",
    )
    parser.add_argument("--generation-pair-samples", type=int, default=1_000_000)
    parser.add_argument("--seed", type=int, default=20_260_802)
    return parser.parse_args()


def import_external_modules(legacy_parser_root: Path, peplink_root: Path) -> tuple[Any, Any]:
    for root in (legacy_parser_root, peplink_root):
        if not root.exists():
            raise FileNotFoundError(root)
        sys.path.insert(0, str(root))
    legacy = importlib.import_module("smiles_to_peptide")
    peplink = importlib.import_module("PepLink")
    if getattr(peplink, "__version__", None) != "0.1.2":
        raise RuntimeError(f"Expected PepLink 0.1.2, found {peplink.__version__}")
    return legacy, peplink


def read_final_peptides(workbook: Path) -> dict[str, dict[str, Any]]:
    if sha256_file(workbook) != EXPECTED_WORKBOOK_SHA256:
        raise RuntimeError("Final-peptide workbook SHA-256 does not match the frozen copy")
    sheet = load_workbook(workbook, data_only=False)["Sheet1"]
    rows: dict[str, dict[str, Any]] = {}
    current_type = ""
    for row_index in range(2, sheet.max_row + 1):
        group = sheet.cell(row_index, 1).value
        if group:
            current_type = str(group).strip()
        apexoracle_id = sheet.cell(row_index, 3).value
        sequence = sheet.cell(row_index, 4).value
        if not apexoracle_id or not str(apexoracle_id).startswith("ApexOracle-"):
            continue
        is_cyclic = current_type.lower() == "cyclic" or bool(
            sheet.cell(row_index, 4).font.underline
        )
        rows[str(apexoracle_id)] = {
            "sequence": str(sequence).strip(),
            "is_cyclic": is_cyclic,
            "workbook_row": row_index,
        }
    if len(rows) != 24:
        raise RuntimeError(f"Expected 24 final peptides, found {len(rows)}")
    return rows


def source_files(directory: Path) -> list[Path]:
    files = [path for path in directory.glob("strain_BAA-*_MIC_1_length_*_noise.txt")]
    parsed = []
    for path in files:
        match = STRICT_PATTERN.match(path.name)
        if match is not None and match.group(1) in TARGET_STRAINS:
            parsed.append(path)
    return sorted(
        parsed,
        key=lambda path: (
            STRICT_PATTERN.match(path.name).group(1),
            int(STRICT_PATTERN.match(path.name).group(2)),
        ),
    )


def build_final_smiles(peplink: Any, sequence: str, is_cyclic: bool) -> tuple[str, str]:
    kwargs: dict[str, Any] = {}
    if is_cyclic:
        kwargs["intrachain_bonds"] = [
            {
                "position1": 1,
                "position2": len(sequence),
                "type": "AMD",
                "chain_participating": "MMB",
            }
        ]
    smiles = peplink.aa_seqs_to_smiles(sequence=sequence, **kwargs)
    selfies = peplink.aa_seqs_to_smiles(
        sequence=sequence, output_format="selfies", **kwargs
    )
    parsed = peplink.smiles_to_aa_seqs(smiles)
    if parsed.sequence is None or bool(parsed.is_cyclic) != is_cyclic:
        raise RuntimeError(f"PepLink failed round-trip validation for {sequence}")
    return canonical_isomeric_smiles(smiles), selfies


def build_corrected_candidates(
    strict_dir: Path,
    final_peptides: dict[str, dict[str, Any]],
    legacy: Any,
    peplink: Any,
    output_dir: Path,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[Path]]:
    mirror_dir = output_dir / "canonical_candidates" / "selfies_files"
    if mirror_dir.exists():
        shutil.rmtree(mirror_dir)
    mirror_dir.mkdir(parents=True, exist_ok=True)

    records: list[dict[str, Any]] = []
    replacements: list[dict[str, Any]] = []
    copied_files: list[Path] = []
    per_rule = Counter()
    global_row = 0

    for source_path in source_files(strict_dir):
        match = STRICT_PATTERN.match(source_path.name)
        assert match is not None
        strain, target_length = match.group(1), int(match.group(2))
        output_lines: list[str] = []
        source_lines = source_path.read_text(encoding="utf-8").splitlines()
        for source_line, source_selfies in enumerate(source_lines, start=1):
            if not source_selfies.strip():
                continue
            global_row += 1
            decoded_smiles = sf.decoder(source_selfies.strip())
            parsed = legacy.smiles_to_pepseq(decoded_smiles)
            if parsed is None or parsed[1] is None:
                raise RuntimeError(f"Legacy parser failed for {source_path}:{source_line}")
            original_sequence = str(parsed[1])
            rule = match_replacement_rule(original_sequence)
            original_canonical = canonical_isomeric_smiles(decoded_smiles)

            if rule is None:
                corrected_sequence = normalize_legacy_arg(original_sequence)
                corrected_selfies = source_selfies.strip()
                corrected_canonical = original_canonical
                apexoracle_id = ""
                replacement_status = "unchanged_generated_output"
            else:
                final_row = final_peptides[rule.apexoracle_id]
                corrected_sequence = str(final_row["sequence"])
                corrected_canonical, corrected_selfies = build_final_smiles(
                    peplink, corrected_sequence, bool(final_row["is_cyclic"])
                )
                if final_row["is_cyclic"]:
                    corrected_sequence = "cyclo-" + corrected_sequence
                apexoracle_id = rule.apexoracle_id
                replacement_status = "nearest_precursor_replaced_by_final_peptide"
                per_rule[rule.apexoracle_id] += 1

            output_lines.append(corrected_selfies)
            record = {
                "candidate_row": global_row,
                "strain": strain,
                "target_mic": 1,
                "target_length": target_length,
                "source_file": source_path.name,
                "source_line": source_line,
                "source_file_sha256": sha256_file(source_path),
                "original_sequence": original_sequence,
                "original_sequence_arg_normalized": normalize_legacy_arg(original_sequence),
                "corrected_sequence": corrected_sequence,
                "apexoracle_id": apexoracle_id,
                "replacement_status": replacement_status,
                "original_selfies": source_selfies.strip(),
                "corrected_selfies": corrected_selfies,
                "original_canonical_isomeric_smiles": original_canonical,
                "corrected_canonical_isomeric_smiles": corrected_canonical,
                "structure_changed": original_canonical != corrected_canonical,
            }
            records.append(record)
            if rule is not None:
                replacements.append(record.copy())

        output_path = mirror_dir / source_path.name
        output_path.write_text(
            "\n".join(output_lines) + ("\n" if output_lines else ""), encoding="utf-8"
        )
        copied_files.append(output_path)

    if len(records) != EXPECTED_STRICT_ROWS:
        raise RuntimeError(f"Expected {EXPECTED_STRICT_ROWS} strict rows, found {len(records)}")
    for rule in REPLACEMENT_RULES:
        if per_rule[rule.apexoracle_id] != rule.expected_occurrences:
            raise RuntimeError(
                f"{rule.apexoracle_id}: expected {rule.expected_occurrences} replacements, "
                f"found {per_rule[rule.apexoracle_id]}"
            )
    return records, replacements, copied_files


def map_final_peptides_to_corrected_candidates(
    final_peptides: dict[str, dict[str, Any]],
    candidates: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for apexoracle_id in sorted(
        final_peptides, key=lambda value: int(value.rsplit("-", 1)[1])
    ):
        final_row = final_peptides[apexoracle_id]
        final_sequence = str(final_row["sequence"])
        topology_sequence = (
            "cyclo-" + final_sequence if final_row["is_cyclic"] else final_sequence
        )
        matches = [
            candidate
            for candidate in candidates
            if topology_aware_equal(
                str(candidate["corrected_sequence"]), topology_sequence
            )
        ]
        if not matches:
            raise RuntimeError(
                f"Final peptide {apexoracle_id} is absent from peptide candidate copy"
            )
        rows.append(
            {
                "apexoracle_id": apexoracle_id,
                "final_sequence": final_sequence,
                "is_cyclic": bool(final_row["is_cyclic"]),
                "matching_candidate_rows": ";".join(
                    str(row["candidate_row"]) for row in matches
                ),
                "matching_occurrences": len(matches),
                "mapping_status": "present_in_corrected_candidate_copy",
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    selected = fieldnames or list(rows[0])
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=selected)
        writer.writeheader()
        writer.writerows(rows)


def read_generation_records(generation_dir: Path) -> tuple[list[dict[str, Any]], list[Path]]:
    rows: list[dict[str, Any]] = []
    files = source_files(generation_dir)
    for source_path in files:
        match = STRICT_PATTERN.match(source_path.name)
        assert match is not None
        strain, target_length = match.group(1), int(match.group(2))
        for source_line, source_selfies in enumerate(
            source_path.read_text(encoding="utf-8").splitlines(), start=1
        ):
            if not source_selfies.strip():
                continue
            decoded_smiles = sf.decoder(source_selfies.strip())
            mol = Chem.MolFromSmiles(decoded_smiles)
            if mol is None:
                raise RuntimeError(f"RDKit failed for {source_path}:{source_line}")
            rows.append(
                {
                    "strain": strain,
                    "target_length": target_length,
                    "source_file": source_path.name,
                    "source_line": source_line,
                    "canonical_isomeric_smiles": Chem.MolToSmiles(
                        mol, canonical=True, isomericSmiles=True
                    ),
                    "_mol": mol,
                }
            )
    if len(rows) != EXPECTED_GENERATION_ROWS:
        raise RuntimeError(
            f"Expected {EXPECTED_GENERATION_ROWS} generation rows, found {len(rows)}"
        )
    return rows, files


def diversity_summary(rows: list[dict[str, Any]], smiles_key: str) -> dict[str, Any]:
    smiles = [str(row[smiles_key]) for row in rows]
    by_strain = Counter(str(row["strain"]) for row in rows)
    unique_by_strain = {
        strain: len(
            {
                str(row[smiles_key])
                for row in rows
                if str(row["strain"]) == strain
            }
        )
        for strain in sorted(by_strain)
    }
    return {
        "saved_rows": len(rows),
        "unique_canonical_isomeric_structures": len(set(smiles)),
        "duplicate_rows_beyond_first": len(rows) - len(set(smiles)),
        "unique_fraction": len(set(smiles)) / len(rows),
        "rows_by_strain": dict(by_strain),
        "unique_structures_by_strain": unique_by_strain,
    }


def add_dataset(hist_rows: list[dict[str, Any]], dataset: str) -> list[dict[str, Any]]:
    return [{"dataset": dataset, **row} for row in hist_rows]


def load_exact_generation_histogram(path: Path) -> list[dict[str, Any]]:
    with path.open(newline="", encoding="utf-8") as handle:
        fine_rows = list(csv.DictReader(handle))
    if len(fine_rows) != 1_000:
        raise RuntimeError(f"Expected 1,000 exact histogram bins, found {len(fine_rows)}")
    counts = np.asarray([int(row["count"]) for row in fine_rows], dtype=np.int64)
    if counts.size % 20:
        raise RuntimeError("Exact histogram cannot be aggregated into 20 bins")
    grouped = counts.reshape(20, counts.size // 20).sum(axis=1)
    total = int(np.sum(grouped))
    expected = EXPECTED_GENERATION_ROWS * (EXPECTED_GENERATION_ROWS - 1) // 2
    if total != expected:
        raise RuntimeError(f"Expected {expected} exact pairs, found {total}")
    rows: list[dict[str, Any]] = []
    for index, count in enumerate(grouped):
        rows.append(
            {
                "bin_left": index / 20,
                "bin_right": (index + 1) / 20,
                "bin_center": (index + 0.5) / 20,
                "count": int(count),
                "fraction": float(count / total),
            }
        )
    return rows


def plot_distributions(
    hist_rows: list[dict[str, Any]],
    output_dir: Path,
) -> list[Path]:
    datasets = ["Peptide candidate pool", "Guided generation outputs"]
    titles = [
        "Peptide candidate pool\n73 candidates",
        "Guided generation outputs\n84,226 outputs",
    ]
    colors = ["#2C6EAA", "#56B4E9"]
    fig, axes = plt.subplots(1, 2, figsize=(9.0, 3.5), sharex=True, sharey=True)
    for axis, dataset, title, color, panel in zip(
        axes, datasets, titles, colors, ("a", "b"), strict=True
    ):
        selected = [row for row in hist_rows if row["dataset"] == dataset]
        centers = np.asarray([row["bin_center"] for row in selected], dtype=float)
        fractions = np.asarray([row["fraction"] for row in selected], dtype=float)
        widths = np.asarray(
            [row["bin_right"] - row["bin_left"] for row in selected], dtype=float
        )
        axis.bar(centers, fractions, width=widths * 0.92, color=color, edgecolor="#333333", linewidth=0.5)
        axis.set_title(title, fontsize=10)
        axis.set_xlabel("Pairwise Tanimoto similarity")
        axis.text(-0.13, 1.03, panel, transform=axis.transAxes, fontweight="bold", fontsize=12)
        axis.grid(axis="y", color="#D9D9D9", linewidth=0.6)
        axis.set_axisbelow(True)
        axis.spines["top"].set_visible(False)
        axis.spines["right"].set_visible(False)
    axes[0].set_ylabel("Fraction of molecule pairs")
    axes[0].set_xlim(0.0, 1.0)
    fig.tight_layout()
    outputs = []
    for suffix in ("png", "pdf", "svg"):
        path = output_dir / f"guided_generation_tanimoto_distributions.{suffix}"
        fig.savefig(path, dpi=300, bbox_inches="tight")
        outputs.append(path)
    plt.close(fig)
    return outputs


def main() -> None:
    args = parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    legacy, peplink = import_external_modules(
        args.legacy_parser_root, args.peplink_root
    )
    final_peptides = read_final_peptides(args.workbook)
    candidates, replacements, copied_files = build_corrected_candidates(
        args.strict_dir,
        final_peptides,
        legacy,
        peplink,
        args.output_dir,
    )

    candidate_records_path = args.output_dir / "canonical_candidates" / "candidates_73.csv"
    replacement_path = args.output_dir / "canonical_candidates" / "replacement_audit.csv"
    write_csv(candidate_records_path, candidates)
    write_csv(replacement_path, replacements)
    final_mapping = map_final_peptides_to_corrected_candidates(
        final_peptides, candidates
    )
    final_mapping_path = (
        args.output_dir / "canonical_candidates" / "final_peptide_mapping.csv"
    )
    write_csv(final_mapping_path, final_mapping)

    candidate_smiles = [
        str(row["corrected_canonical_isomeric_smiles"]) for row in candidates
    ]
    candidate_fingerprints = morgan_fingerprints(candidate_smiles)
    candidate_left, candidate_right, candidate_values = all_pairwise_tanimoto(
        candidate_fingerprints
    )
    candidate_pair_rows = [
        {
            "left_candidate_row": int(left + 1),
            "right_candidate_row": int(right + 1),
            "tanimoto": float(similarity),
        }
        for left, right, similarity in zip(
            candidate_left, candidate_right, candidate_values, strict=True
        )
    ]
    candidate_pairs_path = args.output_dir / "candidate_pairwise_tanimoto.csv"
    write_csv(candidate_pairs_path, candidate_pair_rows)

    generation_rows, generation_files = read_generation_records(args.generation_dir)
    generation_smiles = [str(row["canonical_isomeric_smiles"]) for row in generation_rows]
    generation_fingerprints = morgan_fingerprints_from_mols(
        [row["_mol"] for row in generation_rows]
    )
    cache_dir = args.output_dir / "local_cache"
    cache_dir.mkdir(parents=True, exist_ok=True)
    fingerprint_width = FINGERPRINT_BITS // 8
    fingerprint_bytes = np.empty(
        (len(generation_fingerprints), fingerprint_width), dtype=np.uint8
    )
    for row_index, fingerprint in enumerate(generation_fingerprints):
        binary = DataStructs.BitVectToBinaryText(fingerprint)
        if len(binary) != fingerprint_width:
            raise RuntimeError("Unexpected RDKit fingerprint binary width")
        fingerprint_bytes[row_index] = np.frombuffer(binary, dtype=np.uint8)
    fingerprint_cache_path = (
        cache_dir / "generation_morgan_r2_2048_chiral.npz"
    )
    np.savez(fingerprint_cache_path, fingerprints=fingerprint_bytes)
    sample_left, sample_right = sample_distinct_ordered_pairs(
        len(generation_fingerprints), args.generation_pair_samples, args.seed
    )
    generation_values = sampled_tanimoto(
        generation_fingerprints, sample_left, sample_right
    )

    exact_histogram_path = (
        args.output_dir / "generation_all_pairs_tanimoto_histogram.csv"
    )
    exact_summary_path = (
        args.output_dir / "generation_all_pairs_tanimoto_summary.json"
    )
    if exact_histogram_path.exists() != exact_summary_path.exists():
        raise RuntimeError("Exact generation histogram and summary must exist together")
    exact_generation_summary = None
    if exact_histogram_path.exists():
        generation_histogram = load_exact_generation_histogram(exact_histogram_path)
        exact_generation_summary = json.loads(exact_summary_path.read_text())
    else:
        generation_histogram = histogram_rows(generation_values)
    hist_rows = add_dataset(
        histogram_rows(candidate_values), "Peptide candidate pool"
    ) + add_dataset(generation_histogram, "Guided generation outputs")
    hist_path = args.output_dir / "tanimoto_histogram_plotted_data.csv"
    write_csv(hist_path, hist_rows)
    figure_paths = plot_distributions(hist_rows, args.output_dir)

    candidate_identity = diversity_summary(
        candidates, "corrected_canonical_isomeric_smiles"
    )
    generation_identity = diversity_summary(
        generation_rows, "canonical_isomeric_smiles"
    )
    summary = {
        "protocol": {
            "candidate_level": (
                "All 73 peptide candidate rows; all 2,628 unordered non-self "
                "pairs, including exact duplicate rows."
            ),
            "generation_level": (
                "All pooled target lengths for BAA-3170 and BAA-3197 from the two "
                "MIC-guided output sets; exact canonical-structure uniqueness and all "
                "unordered non-self molecule pairs. The deterministic pair sample is "
                "retained only as an internal stability check."
            ),
            "fingerprint": {
                "type": "RDKit Morgan bit fingerprint",
                "radius": FINGERPRINT_RADIUS,
                "n_bits": FINGERPRINT_BITS,
                "include_chirality": FINGERPRINT_INCLUDE_CHIRALITY,
                "similarity": "Tanimoto",
            },
            "generation_pair_sample_size": args.generation_pair_samples,
            "generation_pair_sample_seed": args.seed,
        },
        "candidate_level": {
            **candidate_identity,
            "replacement_rows": len(replacements),
            "replacement_final_peptides": len(REPLACEMENT_RULES),
            "synthesized_final_peptides_present": len(final_mapping),
            "pairwise_tanimoto": summarize_similarities(candidate_values),
        },
        "generation_level": {
            **generation_identity,
            "pairwise_tanimoto_sample": summarize_similarities(generation_values),
            "pairwise_tanimoto_exact_all_pairs": exact_generation_summary,
        },
    }
    summary_path = args.output_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    input_files = [args.workbook, *source_files(args.strict_dir), *generation_files]
    if exact_generation_summary is not None:
        input_files.extend([exact_histogram_path, exact_summary_path])
    output_files = [
        candidate_records_path,
        replacement_path,
        final_mapping_path,
        candidate_pairs_path,
        hist_path,
        summary_path,
        fingerprint_cache_path,
        *copied_files,
        *figure_paths,
    ]
    manifest = {
        "created_at_utc": datetime.now(timezone.utc).isoformat(),
        "script": str(Path(__file__).resolve()),
        "script_sha256": sha256_file(Path(__file__).resolve()),
        "inputs": [
            {"path": str(path.resolve()), "size": path.stat().st_size, "sha256": sha256_file(path)}
            for path in input_files
        ],
        "outputs": [
            {"path": str(path.resolve()), "size": path.stat().st_size, "sha256": sha256_file(path)}
            for path in output_files
        ],
        "external_code": {
            "legacy_parser": str((args.legacy_parser_root / "smiles_to_peptide.py").resolve()),
            "legacy_parser_sha256": sha256_file(args.legacy_parser_root / "smiles_to_peptide.py"),
            "peplink_root": str(args.peplink_root.resolve()),
            "peplink_version": peplink.__version__,
        },
    }
    manifest_path = args.output_dir / "manifest.json"
    manifest_path.write_text(json.dumps(manifest, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
