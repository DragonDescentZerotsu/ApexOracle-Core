#!/usr/bin/env python3
"""Audit in-house assay species against the MIC guidance training exposure.

The private workbook is read only. Outputs contain assay headers and aggregate
counts, never peptide sequences or row-level activity values.
"""

from __future__ import annotations

import argparse
from collections import Counter
from contextlib import redirect_stdout
import hashlib
from io import StringIO
import json
import os
from pathlib import Path
import re
import sys
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(REPO_ROOT / "src"))

from apexoracle.data.hierarchical_mic_preparation import (  # noqa: E402
    prepare_hierarchical_mic_data,
)


DEFAULT_WORKBOOK = (
    REPO_ROOT
    / "DataPrepare"
    / "Data"
    / "private_inhouse_amp"
    / "Master_List_Peptides_Antimicrobial_Activity.xlsx"
)
DEFAULT_OUTPUT_DIR = DEFAULT_WORKBOOK.parent
DEFAULT_GUIDANCE_PRODUCER = Path(
    os.environ.get(
        "MDLM_GUIDANCE_PRODUCER",
        REPO_ROOT.parent / "mdlm" / "guaidance_regressor_all_data_pad_no_mask.py",
    )
)

# Rules are ordered only for readability. Exactly one must match every assay
# header; this is asserted below so workbook schema drift fails loudly.
HEADER_SPECIES_RULES = (
    ("a. baumannii", "Acinetobacter baumannii"),
    ("e. coli", "Escherichia coli"),
    ("k. pneumoniae", "Klebsiella pneumoniae"),
    ("k. pnemoniae", "Klebsiella pneumoniae"),
    ("p. aeruginosa", "Pseudomonas aeruginosa"),
    ("s. aureus", "Staphylococcus aureus"),
    ("e. faecalis", "Enterococcus faecalis"),
    ("e. faecium", "Enterococcus faecium"),
    ("b. subtilis", "Bacillus subtilis"),
    ("salmonella enterica", "Salmonella enterica"),
    ("s. enterica", "Salmonella enterica"),
    ("l. monocytogenes", "Listeria monocytogenes"),
    ("neisseria meningitis", "Neisseria meningitidis"),
    ("porphyromoras gingivalis", "Porphyromonas gingivalis"),
    ("a. muciniphila", "Akkermansia muciniphila"),
    ("b. eggerthi", "Bacteroides eggerthii"),
    ("b. fragilis", "Bacteroides fragilis"),
    ("b. ovatus", "Bacteroides ovatus"),
    ("b. thetaiotaomicron", "Bacteroides thetaiotaomicron"),
    ("b. uniformis", "Bacteroides uniformis"),
    ("b. vulgatus", "Phocaeicola vulgatus"),
    ("c. aerofaciens", "Collinsella aerofaciens"),
    ("c. scindens", "[Clostridium] scindens"),
    ("c. spiroforme", "Thomasclavelia spiroformis"),
    ("c. spiriforme", "Thomasclavelia spiroformis"),
    ("c. symbiosum", "[Clostridium] symbiosum"),
    ("e. rectale", "Agathobacter rectalis"),
    ("p. copri", "Segatella copri"),
    ("p. distasonis", "Parabacteroides distasonis"),
    ("r. obeum", "Blautia obeum"),
    ("r. torques", "Ruminococcus torques"),
    ("m. luteus", "Micrococcus luteus"),
    ("s. capitis", "Staphylococcus capitis"),
    ("s. epidermidis", "Staphylococcus epidermidis"),
    ("s. haemolyticus", "Staphylococcus haemolyticus"),
    ("s. hominis", "Staphylococcus hominis"),
    ("s. warneri", "Staphylococcus warneri"),
    ("e. cloacae", "Enterobacter cloacae"),
)

# NCBI Datasets v2 was queried on 2026-07-20. These aliases are restricted to
# names present in the workbook and prevent false unseen calls after renaming.
WORKBOOK_TAXONOMY = {
    "Akkermansia muciniphila": (239935, "Akkermansia muciniphila"),
    "Bacteroides eggerthii": (28111, "Bacteroides eggerthii"),
    "Bacteroides ovatus": (28116, "Bacteroides ovatus"),
    "Bacteroides uniformis": (820, "Bacteroides uniformis"),
    "Collinsella aerofaciens": (74426, "Collinsella aerofaciens"),
    "[Clostridium] scindens": (29347, "[Clostridium] scindens"),
    "[Clostridium] symbiosum": (1512, "[Clostridium] symbiosum"),
    "Agathobacter rectalis": (39491, "Agathobacter rectalis"),
    "Segatella copri": (165179, "Segatella copri"),
    "Parabacteroides distasonis": (823, "Parabacteroides distasonis"),
    "Staphylococcus capitis": (29388, "Staphylococcus capitis"),
    "Phocaeicola vulgatus": (821, "Phocaeicola vulgatus"),
    "Blautia obeum": (40520, "Blautia obeum"),
    "Thomasclavelia spiroformis": (29348, "Thomasclavelia spiroformis"),
}

# Exact in-house targets. ``None`` means that the workbook supplies only a
# species label, not a strain identifier suitable for a generation manifest.
TARGET_ACCESSIONS = {
    "Akkermansia muciniphila": "BAA-835",
    "Bacteroides eggerthii": "27754",
    "Bacteroides ovatus": "8483",
    "Bacteroides uniformis": "8492",
    "Collinsella aerofaciens": "25986",
    "[Clostridium] scindens": "35704",
    "[Clostridium] symbiosum": None,
    "Agathobacter rectalis": "33656",
    "Segatella copri": "18205",
    "Parabacteroides distasonis": "8503",
    "Staphylococcus capitis": None,
}

ASSET_SPECIES_ALIASES = {
    "[Clostridium] scindens": ("Clostridium scindens",),
    "[Clostridium] symbiosum": ("Clostridium symbiosum",),
    "Agathobacter rectalis": ("Agathobacter rectalis", "Eubacterium rectale"),
    "Segatella copri": ("Segatella copri", "Prevotella copri"),
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_text(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip().lower()


def map_header_species(header: object) -> str:
    normalized = normalize_text(header)
    matches = [species for token, species in HEADER_SPECIES_RULES if token in normalized]
    if len(matches) != 1:
        raise ValueError(f"Expected one species mapping for {header!r}, got {matches}")
    return matches[0]


def parse_mic(value: object) -> float | None:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value)
    if isinstance(value, str) and re.fullmatch(
        r"[<>]?\s*\d+(?:\.\d+)?", value.strip()
    ):
        return float(re.sub(r"[^0-9.]", "", value))
    return None


def profile_workbook(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        values = list(worksheet.iter_rows(values_only=True))
        headers = values[0]
        for column_index, header in enumerate(headers):
            if column_index < 5 or header is None:
                continue
            if "maximun concentration" in normalize_text(header):
                continue
            measured = [
                parsed
                for row in values[1:]
                if (parsed := parse_mic(row[column_index])) is not None
            ]
            species = map_header_species(header)
            tax_id, ncbi_name = WORKBOOK_TAXONOMY.get(species, (None, species))
            rows.append(
                {
                    "sheet": worksheet.title,
                    "column": column_index + 1,
                    "assay_header": str(header),
                    "species": species,
                    "ncbi_tax_id": tax_id,
                    "ncbi_name": ncbi_name,
                    "measured_cells": len(measured),
                    "minimum_reported_mic": min(measured) if measured else None,
                    "maximum_reported_mic": max(measured) if measured else None,
                }
            )
    table = pd.DataFrame(rows)
    if len(table) != 76:
        raise AssertionError(f"Expected 76 assay headers, found {len(table)}")
    return table


def canonicalize_training_species(
    species_names: Iterable[str], taxonomy_aliases: dict[str, str]
) -> set[str]:
    canonical = set(map(str, species_names))
    for name in list(canonical):
        alias = taxonomy_aliases.get(name)
        if alias is not None:
            canonical.add(alias)
    # Relevant NCBI renamings missing from the frozen producer-era alias file.
    if "Bacteroides vulgatus" in canonical:
        canonical.add("Phocaeicola vulgatus")
    if "Ruminococcus obeum" in canonical:
        canonical.add("Blautia obeum")
    if "Eubacterium rectale" in canonical:
        canonical.add("Agathobacter rectalis")
    if "Prevotella copri" in canonical:
        canonical.add("Segatella copri")
    return canonical


def training_exposure(repo_root: Path):
    # The frozen compatibility helper prints its legacy wrong-ATCC audit. Keep
    # this higher-level report bounded while preserving the exact filtering.
    with redirect_stdout(StringIO()):
        prepared = prepare_hierarchical_mic_data(repo_root)
    actual_ids = set(map(str, prepared.genome_or_text_records[:, 1]))

    # Reproduce the all-data producer's single train-group construction and
    # prove that every eligible ID was actually selected for training.
    strain_for_train: list[str] = []
    repeated_species: list[str] = []
    for species_name, corresponding_ids in prepared.species_to_strains.items():
        if species_name in repeated_species:
            continue
        merged_ids = list(corresponding_ids)
        if species_name in prepared.taxonomy_aliases:
            alias = prepared.taxonomy_aliases[species_name]
            repeated_species.append(alias)
            second = prepared.species_to_strains.get(alias)
            if second is not None:
                merged_ids.extend(second)
        strain_for_train.extend(map(str, merged_ids))
    selected_ids = set(strain_for_train) & actual_ids
    if selected_ids != actual_ids:
        raise AssertionError(
            f"All-data train group omitted {len(actual_ids - selected_ids)} eligible IDs"
        )

    species_by_id: dict[str, str] = {}
    for strain_id in actual_ids:
        species = prepared.atcc_id_to_species.get(
            strain_id
        ) or prepared.original_strain_to_species.get(strain_id)
        if species is None:
            raise AssertionError(f"No species mapping for training strain {strain_id!r}")
        species_by_id[strain_id] = species
    counts = Counter(species_by_id.values())
    canonical_species = canonicalize_training_species(
        counts, prepared.taxonomy_aliases
    )
    return prepared, species_by_id, counts, canonical_species


def normalized_asset_name(path: Path) -> str:
    return normalize_text(
        path.name.replace("_", " ").replace("～", " ").replace("^", "/")
    )


def exact_target_assets(repo_root: Path, species: str, accession: str | None) -> dict:
    if accession is None:
        return {
            "target_accession": None,
            "genome_embedding_matches": 0,
            "text_embedding_matches": 0,
            "generation_assets_ready": False,
            "asset_note": "workbook does not specify an exact strain",
        }
    aliases = ASSET_SPECIES_ALIASES.get(species, (species,))
    alias_tokens = [normalize_text(alias).replace("[", "").replace("]", "") for alias in aliases]
    accession_token = normalize_text(accession).replace("-", " ")

    def matching_files(folder: Path) -> list[Path]:
        matches = []
        for path in folder.iterdir():
            if not path.is_file():
                continue
            name = normalized_asset_name(path).replace("-", " ")
            if accession_token not in name:
                continue
            if any(all(word in name for word in alias.split()) for alias in alias_tokens):
                matches.append(path)
        return matches

    genome_matches = matching_files(repo_root / "DataPrepare" / "Data" / "Genome_embs")
    text_matches = matching_files(
        repo_root / "DataPrepare" / "Data" / "Text_Description" / "ATCC" / "embeddings"
    ) + matching_files(
        repo_root
        / "DataPrepare"
        / "Data"
        / "Text_Description"
        / "wo_ATCC"
        / "embeddings"
    )
    return {
        "target_accession": accession,
        "genome_embedding_matches": len(genome_matches),
        "text_embedding_matches": len(text_matches),
        "generation_assets_ready": bool(genome_matches and text_matches),
        "asset_note": "" if genome_matches and text_matches else "missing exact target embedding asset(s)",
    }


def build_candidate_table(
    assay_headers: pd.DataFrame,
    training_species: set[str],
    repo_root: Path,
) -> pd.DataFrame:
    grouped = assay_headers.groupby("species", sort=True, dropna=False)
    rows: list[dict[str, object]] = []
    training_genera = {
        name.replace("[", "").replace("]", "").split()[0]
        for name in training_species
        if len(name.split()) >= 2
    }
    for species, frame in grouped:
        if species in training_species:
            continue
        genus = species.replace("[", "").replace("]", "").split()[0]
        accession = TARGET_ACCESSIONS.get(species)
        assets = exact_target_assets(repo_root, species, accession)
        rows.append(
            {
                "species": species,
                "ncbi_tax_id": frame["ncbi_tax_id"].dropna().iloc[0]
                if frame["ncbi_tax_id"].notna().any()
                else None,
                "species_absent_from_actual_training": True,
                "current_genus_absent_from_actual_training": genus not in training_genera,
                "assay_columns": len(frame),
                "measured_cells": int(frame["measured_cells"].sum()),
                "assay_headers": " | ".join(frame["assay_header"].astype(str)),
                **assets,
            }
        )
    result = pd.DataFrame(rows).sort_values(
        ["measured_cells", "species"], ascending=[False, True]
    )
    if len(result) != 11:
        raise AssertionError(f"Expected 11 unseen workbook species, found {len(result)}")
    if int((result["measured_cells"] > 0).sum()) != 10:
        raise AssertionError("Expected 10 unseen species with measured MIC values")
    return result


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    parser.add_argument(
        "--guidance-producer", type=Path, default=DEFAULT_GUIDANCE_PRODUCER
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.expanduser().resolve()
    output_dir = args.output_dir.expanduser().resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    assay_headers = profile_workbook(workbook_path)
    prepared, species_by_id, training_counts, training_species = training_exposure(
        REPO_ROOT
    )
    candidates = build_candidate_table(assay_headers, training_species, REPO_ROOT)

    assay_path = output_dir / "reviewer4_inhouse_assay_headers.csv"
    candidate_path = output_dir / "reviewer4_unseen_species_candidates.csv"
    summary_path = output_dir / "reviewer4_species_coverage_summary.json"
    assay_headers.to_csv(assay_path, index=False)
    candidates.to_csv(candidate_path, index=False)

    training_source = (
        REPO_ROOT
        / "DataPrepare"
        / "Data"
        / "DBAASP_inhouse_AMP_SELFIES_token_MIC_Evo.csv"
    )
    producer = args.guidance_producer.expanduser().resolve()
    checkpoint = (
        REPO_ROOT
        / "Checkpoints"
        / "genome_text_learnable_emb"
        / "guidance_regressor_pad_no_mask"
        / "noise_guidance_best_R2_all_peptide_epoch_100.pth"
    )
    summary = {
        "scope": "reviewer 4 in-house species coverage audit",
        "private_workbook": {
            "path": str(workbook_path),
            "bytes": workbook_path.stat().st_size,
            "sha256": sha256_file(workbook_path),
            "sheets": assay_headers["sheet"].drop_duplicates().tolist(),
            "assay_headers": len(assay_headers),
            "unique_species_after_normalization": int(assay_headers["species"].nunique()),
        },
        "guidance_training": {
            "producer_path": str(producer),
            "producer_sha256": sha256_file(producer),
            "training_source_path": str(training_source),
            "training_source_sha256": sha256_file(training_source),
            "actual_genome_text_rows": len(prepared.genome_text_records),
            "actual_genome_or_text_rows": len(prepared.genome_or_text_records),
            "actual_unique_strain_ids": len(species_by_id),
            "producer_era_species_names": len(training_counts),
            "canonical_species_names_with_relevant_aliases": len(training_species),
            "unmapped_actual_strain_ids": 0,
            "checkpoint_path": str(checkpoint),
            "checkpoint_sha256": sha256_file(checkpoint) if checkpoint.exists() else None,
        },
        "comparison": {
            "unseen_species_in_workbook": len(candidates),
            "unseen_species_with_measured_mic": int((candidates["measured_cells"] > 0).sum()),
            "unseen_current_genera_in_workbook": int(
                candidates["current_genus_absent_from_actual_training"].sum()
            ),
            "candidates_with_ready_exact_target_embeddings": int(
                candidates["generation_assets_ready"].sum()
            ),
            "candidate_species": candidates["species"].tolist(),
        },
        "taxonomy": {
            "ncbi_datasets_checked_on": "2026-07-20",
            "relevant_taxa": {
                name: {"tax_id": tax_id, "organism_name": ncbi_name}
                for name, (tax_id, ncbi_name) in WORKBOOK_TAXONOMY.items()
            },
        },
        "outputs": {
            "assay_headers": str(assay_path),
            "unseen_species_candidates": str(candidate_path),
        },
    }
    summary_path.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, indent=2, ensure_ascii=False))
    print("\nUnseen species candidates:")
    print(
        candidates[
            [
                "species",
                "measured_cells",
                "current_genus_absent_from_actual_training",
                "target_accession",
                "generation_assets_ready",
            ]
        ].to_string(index=False)
    )


if __name__ == "__main__":
    main()
