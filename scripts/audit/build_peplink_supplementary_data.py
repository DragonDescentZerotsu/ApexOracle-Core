#!/usr/bin/env python3
"""Build reviewer-facing Supplementary Data for confirmed local AA errors.

This audit is deliberately source-aware. It includes only confirmed local
residue-definition or residue-template errors that intersect the actual
<=512-token DBAASP genome-or-text MIC pool. Whole-peptide structures obtained
through DBAASP-linked PubChem CIDs, predefined demetallated representations,
unsupported PepLink inputs, polymer proxies, and upstream DBAASP annotation
inconsistencies are outside this table.
"""

from __future__ import annotations

import ast
import hashlib
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rdkit import Chem
from rdkit.Chem import rdMolDescriptors

from apexoracle.data.hierarchical_mic_preparation import prepare_hierarchical_mic_data


REPO_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = REPO_ROOT / "DataPrepare" / "Data"
AUDIT_DIR = REPO_ROOT / "experiments" / "peplink_validation"
OUTPUT_DIR = AUDIT_DIR / "supplementary_data"
PUBCHEM_SNAPSHOT = OUTPUT_DIR / "pubchem_corrected_definitions_20260723.json"


@dataclass(frozen=True)
class Correction:
    code: str
    category: str
    correct_name: str
    correct_smiles: str
    error_explanation: str
    correction_action: str
    evidence_basis: str
    evidence_url: str
    stereochemistry_note: str
    confidence: str = "High"


CORRECTIONS = {
    item.code: item
    for item in [
        Correction(
            "N-TYR",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "3-Nitro-L-tyrosine",
            "C1=CC(=C(C=C1C[C@@H](C(=O)O)N)[N+](=O)[O-])O",
            "The historical structure is nitrophenylalanine and omits the "
            "phenolic hydroxyl of nitrotyrosine (C9H10N2O4 instead of C9H10N2O5).",
            "Replace the historical residue structure with 3-nitro-L-tyrosine.",
            "DBAASP annotation/formula and PubChem CID 65124",
            "https://pubchem.ncbi.nlm.nih.gov/compound/65124",
            "The L alpha configuration is specified in the public reference.",
        ),
        Correction(
            "LYS-C18",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "N6-Oleoyl-L-lysine",
            r"CCCCCCCC/C=C\CCCCCCCC(=O)NCCCC[C@@H](C(=O)O)N",
            "The historical structure attaches the oleoyl group to the alpha "
            "amino group rather than the lysine N6 side-chain amino group and "
            "does not encode the cis-9 double-bond geometry.",
            "Replace with N6-oleoyl-L-lysine and retain cis-9 geometry.",
            "DBAASP annotation and PubChem CID 6436547",
            "https://pubchem.ncbi.nlm.nih.gov/compound/6436547",
            "L alpha configuration and cis-9 oleoyl geometry are specified.",
        ),
        Correction(
            "3-Me-Trp",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "Beta-methyltryptophan",
            "CC(C1=CNC2=CC=CC=C21)[C@@H](C(=O)O)N",
            "The historical structure has an incorrect fused indole-derived "
            "connectivity and formula C12H16N2O2. The methyl group belongs on "
            "the side-chain beta carbon, giving C12H14N2O2.",
            "Replace with beta-methyltryptophan.",
            "DBAASP annotation/formula and PubChem CID 53748860",
            "https://pubchem.ncbi.nlm.nih.gov/compound/53748860",
            "The public record fixes the alpha center; beta stereochemistry is unspecified.",
        ),
        Correction(
            "2-OH-Me-SER",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "Alpha-(hydroxymethyl)serine",
            "C(C(CO)(C(=O)O)N)O",
            "The historical structure converts alpha-hydroxymethylserine into "
            "N-hydroxymethylserine by attaching the hydroxymethyl group to the "
            "amino nitrogen rather than the alpha carbon.",
            "Replace with alpha-(hydroxymethyl)serine.",
            "DBAASP annotation/formula and PubChem CID 439893",
            "https://pubchem.ncbi.nlm.nih.gov/compound/439893",
            "The alpha carbon is achiral because it bears two hydroxymethyl groups.",
        ),
        Correction(
            "NNar",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "N-(2-Guanidinoethyl)glycine",
            "N=C(N)NCCNCC(=O)O",
            "The DBAASP formula C5H12N4O2 requires an N-(2-guanidinoethyl)glycine "
            "connectivity. The historical standardized name and structure use a "
            "guanidinopropyl chain and add one carbon (C6H14N4O2).",
            "Replace with the C5 N-(2-guanidinoethyl)glycine structure.",
            "DBAASP formula/connectivity; NnArg nomenclature cross-check",
            "https://patents.google.com/patent/EP3328877B1/en",
            "No stereocenter is present.",
            "Moderate",
        ),
        Correction(
            "D-3-OH-ASN",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "D-3-Hydroxyasparagine",
            "N[C@H](C(O)C(N)=O)C(=O)O",
            "The historical structure is a C3 diamino-hydroxy acid (C3H8N2O3), "
            "not the C4 hydroxyasparagine bearing the side-chain carboxamide "
            "specified by DBAASP (C4H8N2O4).",
            "Replace with D-3-hydroxyasparagine; leave the beta center unspecified.",
            "DBAASP chemical name and formula",
            "https://dbaasp.org/",
            "The D alpha center is encoded as 2R; the DBAASP annotation does not specify the beta center.",
            "Moderate",
        ),
        Correction(
            "IAA-Cys",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "S-Carbamidomethyl-L-cysteine",
            "C([C@@H](C(=O)O)N)SCC(=O)N",
            "Iodoacetamide alkylation yields the S-CH2-CONH2 carbamidomethyl "
            "adduct. The historical structure uses S-CH2-CH2-CONH2 and therefore "
            "contains one extra methylene (C6 rather than C5).",
            "Replace with S-carbamidomethyl-L-cysteine.",
            "DBAASP annotation, PubChem CID 17754220, and UniMod record 4",
            "https://pubchem.ncbi.nlm.nih.gov/compound/17754220; "
            "https://www.unimod.org/modifications_view.php?editid1=4",
            "The L-cysteine alpha center is encoded; its CIP label is R because sulfur has high priority.",
        ),
        Correction(
            "6F-LEU",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "Hexafluoro-L-leucine",
            "C([C@@H](C(=O)O)N)C(C(F)(F)F)C(F)(F)F",
            "The historical structure is a C5 pentafluoro amino acid "
            "(C5H6F5NO2), whereas the DBAASP residue is C6 hexafluoroleucine "
            "(C6H7F6NO2).",
            "Replace with hexafluoro-L-leucine.",
            "DBAASP annotation/formula and PubChem CID 7019591",
            "https://pubchem.ncbi.nlm.nih.gov/compound/7019591",
            "The L alpha configuration is specified in the corrected structure.",
        ),
        Correction(
            "HCha",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "2-Amino-4-cyclohexylbutanoic acid (homocyclohexylalanine)",
            "C1CCC(CC1)CCC(C(=O)O)N",
            "The historical structure is 2-amino-5-cyclohexylpentanoic acid and "
            "contains one extra methylene (C11 instead of C10). The hydrogen "
            "count in the DBAASP text also appears one short for the neutral amino acid.",
            "Replace with the C10 homocyclohexylalanine connectivity.",
            "DBAASP annotation and PubChem CID 224400",
            "https://pubchem.ncbi.nlm.nih.gov/compound/224400",
            "The source record does not specify an L/D alpha configuration.",
        ),
        Correction(
            "D-Me-Trp",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "N-alpha-Methyl-D-tryptophan",
            "CN[C@H](CC1=CNC2=CC=CC=C21)C(=O)O",
            "The historical structure methylates the indole N1 atom. The DBAASP "
            "label denotes methylation of the alpha amino nitrogen.",
            "Replace with N-alpha-methyl-D-tryptophan.",
            "DBAASP annotation/formula and PubChem CID 688254",
            "https://pubchem.ncbi.nlm.nih.gov/compound/688254",
            "The D alpha configuration is encoded as 2R.",
        ),
        Correction(
            "BisHomo-Pra",
            "Incorrect ChatGPT-o1/OPSIN-derived residue definition",
            "2-Aminohept-6-ynoic acid (bishomopropargylglycine)",
            "C#CCCCC(C(=O)O)N",
            "The historical 2-aminohex-5-ynoic-acid structure contains one "
            "methylene too few (C6H9NO2 rather than C7H11NO2).",
            "Replace with 2-aminohept-6-ynoic acid.",
            "DBAASP annotation/formula and PubChem CID 71435968",
            "https://pubchem.ncbi.nlm.nih.gov/compound/71435968",
            "The source record does not specify an L/D alpha configuration.",
        ),
        Correction(
            "Aic",
            "Incorrect or non-polymerizable local residue template",
            "2-Aminoindane-2-carboxylic acid",
            "C1C2=CC=CC=C2CC1(C(=O)O)N",
            "The historical mapping is an aminoguanidine carboxamide heterocycle "
            "and is unrelated to the DBAASP aminoindane carboxylic acid.",
            "Replace with 2-aminoindane-2-carboxylic acid and regenerate the peptide.",
            "DBAASP annotation/formula and PubChem CID 250936",
            "https://pubchem.ncbi.nlm.nih.gov/compound/250936",
            "The alpha carbon is achiral because the two indane ring paths are symmetry-equivalent.",
        ),
        Correction(
            "Agb",
            "Incorrect or non-polymerizable local residue template",
            "2-Amino-4-guanidinobutyric acid (norarginine)",
            "C(CN=C(N)N)C(C(=O)O)N",
            "The historical mapping is a large adamantane/aryl guanidine "
            "compound and is unrelated to the C5 norarginine residue specified by DBAASP.",
            "Replace with 2-amino-4-guanidinobutyric acid and regenerate the peptide.",
            "DBAASP annotation/formula and PubChem CID 435719",
            "https://pubchem.ncbi.nlm.nih.gov/compound/435719",
            "The source record does not specify an L/D alpha configuration.",
        ),
        Correction(
            "Nae",
            "Incorrect or non-polymerizable local residue template",
            "N-(2-Aminoethyl)glycine",
            "C(CNCC(=O)O)N",
            "The historical mapping is a C12 acyl ethanolamide, not the C4 "
            "N-(2-aminoethyl)glycine named and formulated by DBAASP.",
            "Replace the chemical definition and implement explicit peptoid/PNA-backbone connection semantics; exclude until supported.",
            "DBAASP annotation/formula and PubChem CID 428913",
            "https://pubchem.ncbi.nlm.nih.gov/compound/428913",
            "No stereocenter is present; peptide incorporation requires non-alpha-amino-acid attachment rules.",
        ),
        Correction(
            "MIM",
            "Incorrect or non-polymerizable local residue template",
            "L-Mimosine",
            "C1=CN(C=C(C1=O)O)C[C@@H](C(=O)O)N",
            "The historical mapping is a large peptide-like fragment and is "
            "unrelated to the C8 mimosine amino acid specified by DBAASP.",
            "Replace with L-mimosine and regenerate the peptide.",
            "DBAASP annotation/formula and PubChem CID 440473",
            "https://pubchem.ncbi.nlm.nih.gov/compound/440473",
            "The L alpha configuration is specified in the public reference.",
        ),
        Correction(
            "Cl-Th2CA",
            "Incorrect or non-polymerizable local residue template",
            "5-Chlorothiophene-2-carboxylic acid (N-terminal acyl cap)",
            "C1=C(SC(=C1)Cl)C(=O)O",
            "The historical mapping is an unrelated oxirane amide. The DBAASP "
            "entity is a carboxylic acid used at peptide position 1 and should "
            "be represented as a 5-chlorothiophene-2-carbonyl N-terminal cap, "
            "not as an alpha-amino-acid monomer.",
            "Move this definition from the residue map to an N-terminal acyl-cap rule and regenerate the peptide.",
            "DBAASP annotation/formula and PubChem CID 95048",
            "https://pubchem.ncbi.nlm.nih.gov/compound/95048",
            "The free acid is achiral; peptide attachment occurs through its carbonyl as an N-terminal cap.",
        ),
        Correction(
            "D-End",
            "Incorrect secondary ChatGPT-o1/OPSIN-derived residue definition",
            "D-Enduracididine [(2R,4R)-enduracididine]",
            "C1[C@H](NC(=N1)N)C[C@H](C(=O)O)N",
            "The historical mapping is a linear arginine-like residue and omits "
            "the cyclic guanidino side chain of enduracididine.",
            "Replace with the cyclic (2R,4R) D-enduracididine residue.",
            "DBAASP annotation/formula; PubChem CID 15284838 L reference; "
            "J. Nat. Prod. 2018 stereochemical assignment",
            "https://pubchem.ncbi.nlm.nih.gov/compound/15284838; "
            "https://pubs.acs.org/doi/10.1021/acs.jnatprod.8b00354",
            "D-enduracididine is encoded as 2R,4R; the PubChem CID is the 2S,4R L reference used for cross-checking.",
        ),
        Correction(
            "D-IGln",
            "Incorrect secondary ChatGPT-o1/OPSIN-derived residue definition",
            "D-Isoglutamine",
            "C(CC(=O)O)[C@H](C(=O)N)N",
            "The historical structure is a C4 D-asparagine skeleton. "
            "D-isoglutamine is C5H10N2O3 and has the alpha carboxamide and "
            "terminal side-chain carboxyl connectivity shown in the corrected structure.",
            "Replace with D-isoglutamine.",
            "DBAASP annotation/formula and PubChem CID 5288447",
            "https://pubchem.ncbi.nlm.nih.gov/compound/5288447",
            "The D alpha configuration is encoded as 2R.",
        ),
    ]
}

PUBCHEM_CIDS = {
    "N-TYR": 65124,
    "LYS-C18": 6436547,
    "3-Me-Trp": 53748860,
    "2-OH-Me-SER": 439893,
    "IAA-Cys": 17754220,
    "6F-LEU": 7019591,
    "HCha": 224400,
    "D-Me-Trp": 688254,
    "BisHomo-Pra": 71435968,
    "Aic": 250936,
    "Agb": 435719,
    "Nae": 428913,
    "MIM": 440473,
    "Cl-Th2CA": 95048,
    "D-End": 15284838,
    "D-IGln": 5288447,
}


EN_COLUMNS = [
    "DBAASP peptide ID",
    "DBAASP peptide name",
    "DBAASP sequence",
    "Erroneous residue code",
    "Residue position(s)",
    "Erroneous residue occurrences",
    "Eligible model MIC records",
    "Error category",
    "Original DBAASP definition",
    "Historical erroneous SMILES",
    "Historical erroneous formula",
    "Corrected residue/cap",
    "Corrected isomeric SMILES",
    "Corrected formula",
    "What was wrong",
    "Correction required",
    "Stereochemistry/attachment note",
    "Evidence basis",
    "Evidence URL",
    "Confidence",
    "Structure provenance in frozen model data",
    "DBAASP record URL",
]

ZH_COLUMN_MAP = {
    "DBAASP peptide ID": "DBAASP肽ID",
    "DBAASP peptide name": "DBAASP肽名称",
    "DBAASP sequence": "DBAASP序列",
    "Erroneous residue code": "错误残基代码",
    "Residue position(s)": "残基位置",
    "Erroneous residue occurrences": "错误残基出现次数",
    "Eligible model MIC records": "进入模型范围的MIC记录数",
    "Error category": "错误类别",
    "Original DBAASP definition": "DBAASP原始定义",
    "Historical erroneous SMILES": "历史错误SMILES",
    "Historical erroneous formula": "历史错误分子式",
    "Corrected residue/cap": "正确残基或端基名称",
    "Corrected isomeric SMILES": "正确异构SMILES",
    "Corrected formula": "正确分子式",
    "What was wrong": "具体错误",
    "Correction required": "所需修正",
    "Stereochemistry/attachment note": "立体化学或连接方式说明",
    "Evidence basis": "核验依据",
    "Evidence URL": "证据链接",
    "Confidence": "置信度",
    "Structure provenance in frozen model data": "冻结模型数据中的结构来源",
    "DBAASP record URL": "DBAASP记录链接",
}

CATEGORY_ZH = {
    "Incorrect ChatGPT-o1/OPSIN-derived residue definition": "ChatGPT-o1/OPSIN派生残基定义错误",
    "Incorrect or non-polymerizable local residue template": "本地残基模板错误或不可直接聚合",
    "Incorrect secondary ChatGPT-o1/OPSIN-derived residue definition": "二次ChatGPT-o1/OPSIN支路残基定义错误",
}

ZH_DETAILS = {
    "N-TYR": {
        "correct_name": "3-硝基-L-酪氨酸",
        "source_definition": "硝基酪氨酸；DBAASP给出的分子式为C9H10N2O5。",
        "error": "历史结构实际是硝基苯丙氨酸，遗漏了硝基酪氨酸的酚羟基，因此为C9H10N2O4而不是C9H10N2O5。",
        "action": "用3-硝基-L-酪氨酸替换历史残基结构。",
        "stereo": "公开结构记录明确给出L型α位构型。",
        "evidence": "DBAASP名称/分子式与PubChem CID 65124。",
    },
    "LYS-C18": {
        "correct_name": "N6-油酰-L-赖氨酸",
        "source_definition": "侧链被油酰化的赖氨酸。",
        "error": "历史结构把油酰基接在α-氨基上，而不是赖氨酸侧链N6氨基上，同时没有编码cis-9双键几何构型。",
        "action": "替换为N6-油酰-L-赖氨酸，并保留cis-9构型。",
        "stereo": "正确结构明确给出L型α位和cis-9油酰双键构型。",
        "evidence": "DBAASP名称与PubChem CID 6436547。",
    },
    "3-Me-Trp": {
        "correct_name": "β-甲基色氨酸",
        "source_definition": "β-甲基色氨酸（3-甲基色氨酸）；DBAASP分子式为C12H14N2O2。",
        "error": "历史结构具有错误的稠合吲哚连接关系，分子式为C12H16N2O2；甲基应位于侧链β碳，正确分子式为C12H14N2O2。",
        "action": "替换为β-甲基色氨酸。",
        "stereo": "公开结构固定α位构型；β位立体化学未指定。",
        "evidence": "DBAASP名称/分子式与PubChem CID 53748860。",
    },
    "2-OH-Me-SER": {
        "correct_name": "α-(羟甲基)丝氨酸",
        "source_definition": "α-羟甲基丝氨酸；DBAASP分子式为C4H9NO4。",
        "error": "历史结构把羟甲基连接到氨基氮上，得到N-羟甲基丝氨酸；正确连接位点应为α碳。",
        "action": "替换为α-(羟甲基)丝氨酸。",
        "stereo": "α碳连接两个相同的羟甲基，因此不是手性中心。",
        "evidence": "DBAASP名称/分子式与PubChem CID 439893。",
    },
    "NNar": {
        "correct_name": "N-(2-胍基乙基)甘氨酸",
        "source_definition": "N-(胍基乙基)甘氨酸；DBAASP分子式为C5H12N4O2。",
        "error": "DBAASP的C5分子式要求N-(2-胍基乙基)甘氨酸连接关系；历史标准化名称和结构使用胍基丙基链，多出一个碳，成为C6H14N4O2。",
        "action": "替换为C5的N-(2-胍基乙基)甘氨酸结构。",
        "stereo": "该分子没有手性中心。",
        "evidence": "DBAASP分子式/连接关系，并用NnArg命名资料交叉核验。",
    },
    "D-3-OH-ASN": {
        "correct_name": "D-3-羟基天冬酰胺",
        "source_definition": "D-3-羟基天冬酰胺；DBAASP分子式为C4H8N2O4。",
        "error": "历史结构是C3二氨基羟基酸（C3H8N2O3），缺少DBAASP所定义的C4羟基天冬酰胺侧链酰胺。",
        "action": "替换为D-3-羟基天冬酰胺；β位保持未指定。",
        "stereo": "D型α位编码为2R；DBAASP没有指定β位构型。",
        "evidence": "DBAASP化学名称与分子式。",
    },
    "IAA-Cys": {
        "correct_name": "S-氨甲酰甲基-L-半胱氨酸",
        "source_definition": "碘乙酰胺修饰的半胱氨酸。",
        "error": "碘乙酰胺烷基化应产生S-CH2-CONH2氨甲酰甲基加合物；历史结构为S-CH2-CH2-CONH2，多出一个亚甲基（C6而不是C5）。",
        "action": "替换为S-氨甲酰甲基-L-半胱氨酸。",
        "stereo": "编码L-半胱氨酸α位；由于硫原子优先级较高，其CIP标记为R。",
        "evidence": "DBAASP名称、PubChem CID 17754220和UniMod record 4。",
    },
    "6F-LEU": {
        "correct_name": "六氟-L-亮氨酸",
        "source_definition": "六氟亮氨酸；DBAASP分子式为C6H7F6NO2。",
        "error": "历史结构是C5五氟氨基酸（C5H6F5NO2），而DBAASP定义的是C6六氟亮氨酸（C6H7F6NO2）。",
        "action": "替换为六氟-L-亮氨酸。",
        "stereo": "正确结构明确编码L型α位。",
        "evidence": "DBAASP名称/分子式与PubChem CID 7019591。",
    },
    "HCha": {
        "correct_name": "2-氨基-4-环己基丁酸（homocyclohexylalanine）",
        "source_definition": "homocyclohexylalanine；DBAASP标注为C10骨架。",
        "error": "历史结构是2-氨基-5-环己基戊酸，多出一个亚甲基（C11而不是C10）；DBAASP文本中的中性分子氢数似乎也少写了一个。",
        "action": "替换为C10的homocyclohexylalanine连接关系。",
        "stereo": "来源没有指定α位L/D构型。",
        "evidence": "DBAASP名称与PubChem CID 224400。",
    },
    "D-Me-Trp": {
        "correct_name": "Nα-甲基-D-色氨酸",
        "source_definition": "D-N-甲基色氨酸；DBAASP分子式为C12H14N2O2。",
        "error": "历史结构甲基化的是吲哚N1；DBAASP标签表示α-氨基氮甲基化。",
        "action": "替换为Nα-甲基-D-色氨酸。",
        "stereo": "D型α位编码为2R。",
        "evidence": "DBAASP名称/分子式与PubChem CID 688254。",
    },
    "BisHomo-Pra": {
        "correct_name": "2-氨基庚-6-炔酸（bishomopropargylglycine）",
        "source_definition": "bishomopropargylglycine；DBAASP分子式为C7H11NO2。",
        "error": "历史2-氨基己-5-炔酸结构少一个亚甲基，为C6H9NO2而不是C7H11NO2。",
        "action": "替换为2-氨基庚-6-炔酸。",
        "stereo": "来源没有指定α位L/D构型。",
        "evidence": "DBAASP名称/分子式与PubChem CID 71435968。",
    },
    "Aic": {
        "correct_name": "2-氨基茚满-2-羧酸",
        "source_definition": "2-氨基茚满-2-羧酸；DBAASP分子式为C10H11NO2。",
        "error": "历史mapping是氨基胍/甲酰胺杂环，与DBAASP定义的氨基茚满羧酸无关。",
        "action": "替换为2-氨基茚满-2-羧酸并重新生成完整肽。",
        "stereo": "α碳连接的两条茚满环路径对称等价，因此不是手性中心。",
        "evidence": "DBAASP名称/分子式与PubChem CID 250936。",
    },
    "Agb": {
        "correct_name": "2-氨基-4-胍基丁酸（norarginine）",
        "source_definition": "2-氨基-4-胍基丁酸；DBAASP分子式为C5H12N4O2。",
        "error": "历史mapping是大型金刚烷/芳基胍化合物，与DBAASP定义的C5 norarginine无关。",
        "action": "替换为2-氨基-4-胍基丁酸并重新生成完整肽。",
        "stereo": "来源没有指定α位L/D构型。",
        "evidence": "DBAASP名称/分子式与PubChem CID 435719。",
    },
    "Nae": {
        "correct_name": "N-(2-氨基乙基)甘氨酸",
        "source_definition": "N-(2-氨基乙基)甘氨酸；DBAASP分子式为C4H10N2O2。",
        "error": "历史mapping是C12酰基乙醇酰胺，不是DBAASP命名并给出分子式的C4 N-(2-氨基乙基)甘氨酸。",
        "action": "替换自由化合物定义，并实现明确的peptoid/PNA骨架连接语义；在支持前排除。",
        "stereo": "没有手性中心；并入肽链需要非α-氨基酸的连接规则。",
        "evidence": "DBAASP名称/分子式与PubChem CID 428913。",
    },
    "MIM": {
        "correct_name": "L-含羞草氨酸（mimosine）",
        "source_definition": "含羞草氨酸；DBAASP分子式为C8H10N2O4。",
        "error": "历史mapping是大型肽样片段，与DBAASP定义的C8含羞草氨酸无关。",
        "action": "替换为L-含羞草氨酸并重新生成完整肽。",
        "stereo": "公开结构记录明确给出L型α位。",
        "evidence": "DBAASP名称/分子式与PubChem CID 440473。",
    },
    "Cl-Th2CA": {
        "correct_name": "5-氯噻吩-2-羧酸（N端酰基帽）",
        "source_definition": "5-氯噻吩-2-羧酸；DBAASP分子式为C5H3ClO2S。",
        "error": "历史mapping是无关的环氧化物酰胺。DBAASP实体位于肽的第1位，应表示为5-氯噻吩-2-羰基N端酰基帽，而不是α-氨基酸单体。",
        "action": "从residue mapping移到N端酰基帽规则，并重新生成完整肽。",
        "stereo": "游离酸没有手性；并入肽时通过羰基形成N端酰基连接。",
        "evidence": "DBAASP名称/分子式与PubChem CID 95048。",
    },
    "D-End": {
        "correct_name": "D-Enduracididine〔(2R,4R)-enduracididine〕",
        "source_definition": "D-Enduracididine（环状精氨酸类似物）；DBAASP分子式为C6H12N4O2。",
        "error": "历史mapping是线性精氨酸样残基，遗漏了enduracididine的环状胍基侧链。",
        "action": "替换为环状(2R,4R)-D-enduracididine残基。",
        "stereo": "D-enduracididine编码为2R,4R；PubChem CID是用于交叉核验的2S,4R L型参考。",
        "evidence": "DBAASP名称/分子式、PubChem CID 15284838的L型参考及J. Nat. Prod. 2018立体化学指认。",
    },
    "D-IGln": {
        "correct_name": "D-异谷氨酰胺（D-isoglutamine）",
        "source_definition": "D-isoglutamine；DBAASP分子式为C5H10N2O3。",
        "error": "历史结构是C4 D-天冬酰胺骨架；D-isoglutamine应为C5H10N2O3，并具有α位甲酰胺和侧链末端羧基。",
        "action": "替换为D-isoglutamine。",
        "stereo": "D型α位编码为2R。",
        "evidence": "DBAASP名称/分子式与PubChem CID 5288447。",
    },
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def molecular_formula(smiles: str) -> str:
    molecule = Chem.MolFromSmiles(smiles)
    if molecule is None:
        raise ValueError(f"RDKit could not parse SMILES: {smiles}")
    return rdMolDescriptors.CalcMolFormula(molecule)


def canonical_smiles(smiles: str, *, isomeric: bool) -> str:
    molecule = Chem.MolFromSmiles(smiles)
    if molecule is None:
        raise ValueError(f"RDKit could not parse SMILES: {smiles}")
    return Chem.MolToSmiles(molecule, isomericSmiles=isomeric)


def validate_pubchem_snapshot() -> None:
    payload = json.loads(PUBCHEM_SNAPSHOT.read_text())
    records = payload["PropertyTable"]["Properties"]
    by_cid = {int(record["CID"]): record for record in records}
    if len(records) != len(PUBCHEM_CIDS) or len(by_cid) != len(PUBCHEM_CIDS):
        raise ValueError(
            f"Expected {len(PUBCHEM_CIDS)} unique PubChem records; found {len(records)}"
        )
    for code, cid in PUBCHEM_CIDS.items():
        correction = CORRECTIONS[code]
        record = by_cid[cid]
        expected_formula = molecular_formula(correction.correct_smiles)
        if record["MolecularFormula"] != expected_formula:
            raise ValueError(
                f"{code}/CID {cid} formula mismatch: "
                f"{record['MolecularFormula']} != {expected_formula}"
            )
        # PubChem CID 15284838 is the L-enduracididine reference. The corrected
        # D-End entry intentionally inverts only the alpha center, so compare
        # connectivity here and retain the literature-backed 2R,4R assignment.
        isomeric = code != "D-End"
        if canonical_smiles(record["SMILES"], isomeric=isomeric) != canonical_smiles(
            correction.correct_smiles, isomeric=isomeric
        ):
            raise ValueError(f"{code}/CID {cid} structure mismatch")


def unusual_name(item: dict) -> str:
    return item.get("name") or (item.get("modificationType") or {}).get("name") or ""


def unusual_description(item: dict) -> str:
    return (
        item.get("description")
        or (item.get("modificationType") or {}).get("description")
        or ""
    ).strip()


def eligible_model_counts(curated_ids: set[int]) -> tuple[Counter[int], int, int]:
    prepared = prepare_hierarchical_mic_data(REPO_ROOT)
    selected: list[tuple] = []
    for raw_row in prepared.genome_or_text_records:
        row = tuple(raw_row)
        try:
            peptide_id = int(row[0])
            token_count = len(ast.literal_eval(row[2]))
        except (TypeError, ValueError, SyntaxError):
            continue
        if peptide_id in curated_ids and token_count <= 512:
            selected.append(row)
    counts = Counter(int(row[0]) for row in selected)
    return counts, len(counts), len(selected)


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        if sheet.max_row > 1 and sheet.max_column > 1:
            sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for column_cells in sheet.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells[: min(sheet.max_row, 100)]
            )
            width = min(max(max_len + 2, 12), 52)
            sheet.column_dimensions[column_cells[0].column_letter].width = width
        sheet.sheet_view.showGridLines = False
    workbook.save(path)


def translate_frame(frame: pd.DataFrame) -> pd.DataFrame:
    translated = frame.rename(columns=ZH_COLUMN_MAP).copy()
    translated["错误类别"] = translated["错误类别"].map(CATEGORY_ZH)
    translated["置信度"] = translated["置信度"].map(
        {"High": "高", "Moderate": "中"}
    )
    translated["DBAASP原始定义"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["source_definition"]
    )
    translated["正确残基或端基名称"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["correct_name"]
    )
    translated["具体错误"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["error"]
    )
    translated["所需修正"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["action"]
    )
    translated["立体化学或连接方式说明"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["stereo"]
    )
    translated["核验依据"] = translated["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["evidence"]
    )
    translated["冻结模型数据中的结构来源"] = "本地残基拼接（不是整肽PubChem CID结构）"
    return translated


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    validate_pubchem_snapshot()

    scope = pd.read_csv(AUDIT_DIR / "recalculated_local_error_peptides.csv")
    affected_ids = set(scope["DBAASP_id"].astype(int))
    if len(affected_ids) != 56:
        raise ValueError(f"Expected 56 affected peptides, found {len(affected_ids)}")

    records = {
        int(record["id"]): record
        for record in json.loads((DATA_DIR / "all_peptides_data.json").read_text())
    }
    mapping = pd.read_csv(DATA_DIR / "all_aa_smiles_new_handcrafted.csv").set_index(
        "aa"
    )["SMILES"]
    missing_mapping = sorted(set(CORRECTIONS) - set(mapping.index))
    if missing_mapping:
        raise ValueError(f"Missing historical mappings: {missing_mapping}")

    curated = pd.read_csv(DATA_DIR / "DBAASP_id_bact_name_SMILES_MIC_Evo.csv")
    curated_ids = set(curated["DBAASP_id"].astype(int))
    model_counts, pool_peptides, pool_rows = eligible_model_counts(curated_ids)
    if (pool_peptides, pool_rows) != (15177, 74103):
        raise ValueError(
            "Eligible model pool changed: expected 15,177 peptides/74,103 rows, "
            f"found {pool_peptides:,}/{pool_rows:,}"
        )

    rows = []
    for peptide_id in sorted(affected_ids):
        record = records[peptide_id]
        affected_items = [
            item
            for item in (record.get("unusualAminoAcids") or [])
            if unusual_name(item) in CORRECTIONS
        ]
        affected_codes = sorted({unusual_name(item) for item in affected_items})
        if len(affected_codes) != 1:
            raise ValueError(
                f"Expected one erroneous definition for peptide {peptide_id}; "
                f"found {affected_codes}"
            )
        code = affected_codes[0]
        correction = CORRECTIONS[code]
        source_definitions = sorted(
            {unusual_description(item) for item in affected_items if unusual_description(item)}
        )
        positions = sorted(int(item["position"]) for item in affected_items)
        historical_smiles = str(mapping.loc[code])
        correct_formula = molecular_formula(correction.correct_smiles)
        rows.append(
            {
                "DBAASP peptide ID": peptide_id,
                "DBAASP peptide name": record.get("name") or "",
                "DBAASP sequence": record.get("sequence") or "",
                "Erroneous residue code": code,
                "Residue position(s)": ", ".join(map(str, positions)),
                "Erroneous residue occurrences": len(positions),
                "Eligible model MIC records": model_counts[peptide_id],
                "Error category": correction.category,
                "Original DBAASP definition": " | ".join(source_definitions),
                "Historical erroneous SMILES": historical_smiles,
                "Historical erroneous formula": molecular_formula(historical_smiles),
                "Corrected residue/cap": correction.correct_name,
                "Corrected isomeric SMILES": Chem.MolToSmiles(
                    Chem.MolFromSmiles(correction.correct_smiles), isomericSmiles=True
                ),
                "Corrected formula": correct_formula,
                "What was wrong": correction.error_explanation,
                "Correction required": correction.correction_action,
                "Stereochemistry/attachment note": correction.stereochemistry_note,
                "Evidence basis": correction.evidence_basis,
                "Evidence URL": correction.evidence_url,
                "Confidence": correction.confidence,
                "Structure provenance in frozen model data": (
                    "Local residue assembly (not a whole-peptide PubChem CID structure)"
                ),
                "DBAASP record URL": f"https://dbaasp.org/peptide-card?id=DBAASPR_{peptide_id}",
            }
        )

    peptide_frame = pd.DataFrame(rows, columns=EN_COLUMNS)
    if len(peptide_frame) != 56:
        raise ValueError(f"Expected 56 peptide rows, found {len(peptide_frame)}")
    if int(peptide_frame["Eligible model MIC records"].sum()) != 219:
        raise ValueError(
            "Expected 219 affected MIC rows, found "
            f"{peptide_frame['Eligible model MIC records'].sum()}"
        )

    definition_frame = (
        peptide_frame.groupby("Erroneous residue code", sort=True)
        .agg(
            **{
                "Affected peptides": ("DBAASP peptide ID", "nunique"),
                "Eligible model MIC records": ("Eligible model MIC records", "sum"),
                "Error category": ("Error category", "first"),
                "Original DBAASP definition": ("Original DBAASP definition", "first"),
                "Historical erroneous SMILES": ("Historical erroneous SMILES", "first"),
                "Historical erroneous formula": ("Historical erroneous formula", "first"),
                "Corrected residue/cap": ("Corrected residue/cap", "first"),
                "Corrected isomeric SMILES": ("Corrected isomeric SMILES", "first"),
                "Corrected formula": ("Corrected formula", "first"),
                "What was wrong": ("What was wrong", "first"),
                "Correction required": ("Correction required", "first"),
                "Stereochemistry/attachment note": (
                    "Stereochemistry/attachment note",
                    "first",
                ),
                "Evidence basis": ("Evidence basis", "first"),
                "Evidence URL": ("Evidence URL", "first"),
                "Confidence": ("Confidence", "first"),
            }
        )
        .reset_index()
    )
    if len(definition_frame) != 18:
        raise ValueError(f"Expected 18 erroneous definitions, found {len(definition_frame)}")

    readme = pd.DataFrame(
        [
            (
                "Purpose",
                "Reviewer-facing record-level audit of confirmed local amino-acid "
                "definition or residue-template errors that affected the model input pool.",
            ),
            ("Affected scope", "56 of 15,177 DBAASP peptides (0.369%)."),
            ("Affected MIC scope", "219 of 74,103 eligible MIC records (0.296%)."),
            (
                "Eligibility",
                "DBAASP rows in the canonical genome-or-text loader with molecular "
                "token length <=512, matching the model data path. The 74,103-row "
                "denominator retains the loader's modality-path multiplicity.",
            ),
            (
                "Included",
                "Confirmed local ChatGPT-o1/OPSIN-derived residue-definition errors "
                "and confirmed local residue-template errors.",
            ),
            (
                "Excluded",
                "Whole-peptide structures retrieved by DBAASP-linked PubChem CID; "
                "predefined demetallated representations; unsupported PepLink inputs; "
                "polymer proxies; ambiguous source annotations; and upstream DBAASP "
                "sequence/annotation inconsistencies.",
            ),
            (
                "Interpretation",
                "The corrected structures in this file are residue/free-compound or "
                "terminal-cap definitions. They do not imply that all 56 complete "
                "peptide structures have already been regenerated.",
            ),
            (
                "Reproducibility",
                "The original frozen inputs are retained only to reproduce the "
                "reported models. These records are flagged for correction or "
                "exclusion in the successor dataset.",
            ),
            (
                "Manual review",
                "Public structure records were used where available. NNar and "
                "D-3-OH-ASN are explicitly marked Moderate because their corrected "
                "connectivity relies on the DBAASP name/formula rather than an exact "
                "public compound record; unspecified stereocenters remain unspecified.",
            ),
        ],
        columns=["Field", "Description"],
    )

    peptide_csv = OUTPUT_DIR / "Supplementary_Data_affected_peptides.csv"
    definition_csv = OUTPUT_DIR / "Supplementary_Data_residue_definitions.csv"
    peptide_frame.to_csv(peptide_csv, index=False)
    definition_frame.to_csv(definition_csv, index=False)

    english_xlsx = OUTPUT_DIR / "Supplementary_Data_AA_conversion_errors.xlsx"
    with pd.ExcelWriter(english_xlsx, engine="openpyxl") as writer:
        readme.to_excel(writer, sheet_name="README", index=False)
        peptide_frame.to_excel(writer, sheet_name="Affected_peptides", index=False)
        definition_frame.to_excel(writer, sheet_name="Residue_definitions", index=False)
    style_workbook(english_xlsx)

    zh_peptide = translate_frame(peptide_frame)
    zh_definition = definition_frame.rename(
        columns={
            "Erroneous residue code": "错误残基代码",
            "Affected peptides": "受影响肽数",
            "Eligible model MIC records": "进入模型范围的MIC记录数",
            "Error category": "错误类别",
            "Original DBAASP definition": "DBAASP原始定义",
            "Historical erroneous SMILES": "历史错误SMILES",
            "Historical erroneous formula": "历史错误分子式",
            "Corrected residue/cap": "正确残基或端基名称",
            "Corrected isomeric SMILES": "正确异构SMILES",
            "Corrected formula": "正确分子式",
            "What was wrong": "具体错误",
            "Correction required": "所需修正",
            "Stereochemistry/attachment note": "立体化学或连接方式说明",
            "Evidence basis": "核验依据",
            "Evidence URL": "证据链接",
            "Confidence": "置信度",
        }
    )
    zh_definition["错误类别"] = zh_definition["错误类别"].map(CATEGORY_ZH)
    zh_definition["置信度"] = zh_definition["置信度"].map(
        {"High": "高", "Moderate": "中"}
    )
    zh_definition["DBAASP原始定义"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["source_definition"]
    )
    zh_definition["正确残基或端基名称"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["correct_name"]
    )
    zh_definition["具体错误"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["error"]
    )
    zh_definition["所需修正"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["action"]
    )
    zh_definition["立体化学或连接方式说明"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["stereo"]
    )
    zh_definition["核验依据"] = zh_definition["错误残基代码"].map(
        lambda code: ZH_DETAILS[code]["evidence"]
    )
    zh_readme = pd.DataFrame(
        [
            ("用途", "面向审稿人的逐肽审计：仅记录实际影响模型输入范围的已确认本地AA定义或残基模板错误。"),
            ("受影响范围", "15,177个DBAASP肽中的56个（0.369%）。"),
            ("受影响MIC范围", "74,103条合格MIC记录中的219条（0.296%）。"),
            (
                "纳入条件",
                "模型canonical genome-or-text loader中分子token长度不超过512的DBAASP行；"
                "74,103行的分母保留loader中的modality-path重复计数。",
            ),
            ("纳入内容", "已确认的本地ChatGPT-o1/OPSIN派生残基定义错误和本地残基模板错误。"),
            (
                "排除内容",
                "DBAASP链接PubChem CID取得的整肽结构、预定义去金属表示、仅超出PepLink支持范围的输入、"
                "polymer proxy、来源含糊定义，以及DBAASP序列与annotation不一致等上游脏数据。",
            ),
            (
                "如何理解正确版本",
                "表中的正确结构是残基自由化合物或端基定义，不表示56个完整肽结构已经全部重新生成。",
            ),
            (
                "复现边界",
                "原冻结输入仅用于复现已报告模型；后续数据版本应修正这些定义，或在连接语义尚未实现时排除。",
            ),
            (
                "人工核验说明",
                "有公开结构记录时优先使用公开记录。NNar和D-3-OH-ASN仅依据DBAASP名称/分子式确认连接关系，"
                "因此标记为中等置信度；来源未指定的立体中心保持未指定。",
            ),
        ],
        columns=["字段", "说明"],
    )
    zh_csv = OUTPUT_DIR / "补充数据_受影响肽_中文.csv"
    zh_peptide.to_csv(zh_csv, index=False)
    zh_xlsx = OUTPUT_DIR / "补充数据_AA转换错误_中文.xlsx"
    with pd.ExcelWriter(zh_xlsx, engine="openpyxl") as writer:
        zh_readme.to_excel(writer, sheet_name="说明", index=False)
        zh_peptide.to_excel(writer, sheet_name="受影响肽", index=False)
        zh_definition.to_excel(writer, sheet_name="错误残基定义", index=False)
    style_workbook(zh_xlsx)

    source_paths = [
        DATA_DIR / "all_peptides_data.json",
        DATA_DIR / "all_aa_smiles_new_handcrafted.csv",
        DATA_DIR / "DBAASP_id_bact_name_SMILES_MIC_Evo.csv",
        AUDIT_DIR / "recalculated_local_error_peptides.csv",
        PUBCHEM_SNAPSHOT,
    ]
    output_paths = [
        peptide_csv,
        definition_csv,
        english_xlsx,
        zh_csv,
        zh_xlsx,
    ]
    summary = {
        "schema_version": 1,
        "scope": {
            "eligible_peptides": pool_peptides,
            "eligible_mic_records": pool_rows,
            "affected_peptides": int(peptide_frame["DBAASP peptide ID"].nunique()),
            "affected_mic_records": int(
                peptide_frame["Eligible model MIC records"].sum()
            ),
            "affected_residue_definitions": int(
                peptide_frame["Erroneous residue code"].nunique()
            ),
            "peptide_fraction": 56 / 15177,
            "mic_record_fraction": 219 / 74103,
        },
        "scope_exclusions": [
            "whole-peptide PubChem structures",
            "predefined demetallated representations",
            "unsupported PepLink inputs",
            "non-exact polymer proxies",
            "ambiguous source annotations",
            "upstream DBAASP sequence/annotation inconsistencies",
        ],
        "source_sha256": {
            str(path.relative_to(REPO_ROOT)): sha256(path) for path in source_paths
        },
        "output_sha256": {
            str(path.relative_to(REPO_ROOT)): sha256(path) for path in output_paths
        },
    }
    summary_path = OUTPUT_DIR / "supplementary_data_manifest.json"
    summary_path.write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")

    print(f"English workbook: {english_xlsx}")
    print(f"Chinese workbook: {zh_xlsx}")
    print(
        "Validated scope: "
        f"{len(peptide_frame)}/{pool_peptides} peptides; "
        f"{peptide_frame['Eligible model MIC records'].sum()}/{pool_rows} MIC records; "
        f"{len(definition_frame)} definitions"
    )
    print(f"Manifest: {summary_path}")


if __name__ == "__main__":
    main()
